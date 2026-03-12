"""
Funções de exportação de relatórios para Excel e PDF.
"""

import io
from datetime import datetime
from typing import Optional
import pandas as pd

# Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# PDF
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer


def export_to_excel(df: pd.DataFrame, title: str) -> Optional[bytes]:
    """
    Exporta DataFrame para arquivo Excel formatado.

    Args:
        df: DataFrame com os dados
        title: Título do relatório

    Returns:
        Bytes do arquivo Excel ou None em caso de erro
    """
    try:
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_alignment = Alignment(horizontal="left", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Título
        ws.merge_cells("A1:H1")
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        # Data de geração
        ws.merge_cells("A2:H2")
        ws["A2"] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws["A2"].font = Font(italic=True, size=10)
        ws["A2"].alignment = Alignment(horizontal="center")

        # Espaço
        start_row = 4

        # Preparar DataFrame para exportação
        df_export = df.copy()

        # Converter colunas booleanas para texto
        bool_columns = df_export.select_dtypes(include=["bool"]).columns
        for col in bool_columns:
            df_export[col] = df_export[col].apply(lambda x: "Sim" if x else "Não")

        # Escrever dados
        for r_idx, row in enumerate(dataframe_to_rows(df_export, index=False, header=True)):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=start_row + r_idx, column=c_idx, value=value)
                cell.border = thin_border

                if r_idx == 0:  # Header
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                else:
                    cell.alignment = cell_alignment

        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(output)
        return output.getvalue()

    except Exception as e:
        print(f"Erro ao exportar Excel: {e}")
        return None


def export_to_pdf(
    df: pd.DataFrame,
    title: str,
    year: int,
    month: Optional[int] = None,
) -> Optional[bytes]:
    """
    Exporta DataFrame para PDF landscape com quebra de linha correta nas células.
    'Observações' é excluída automaticamente. Inclui resumo geral e por empresa.
    """
    # ── Estilos de célula (Paragraph habilita word-wrap dentro da tabela) ─────
    cell_style = ParagraphStyle(
        "Cell", fontName="Helvetica", fontSize=6.5, leading=8, spaceAfter=0,
    )
    header_style = ParagraphStyle(
        "Header", fontName="Helvetica-Bold", fontSize=6.5, leading=8,
        textColor=colors.whitesmoke, alignment=1, spaceAfter=0,
    )

    # Nomes abreviados para o cabeçalho (evita células muito largas)
    HEADER_ALIASES = {
        "Fibra Lançada (m)":      "Fibra (m)",
        "Abert./Fech. Cx Emenda": "Ab. Cx\nEmenda",
        "Abert./Fech. CTO":       "Ab.\nCTO",
        "Abert./Fech. Rozeta":    "Ab.\nRozeta",
        "Qtd Cx Emenda":          "Cx\nEmenda",
        "Qtd CTOs":               "CTOs",
        "Concluída em":           "Concluída\nem",
    }

    # Larguras fixas por coluna (landscape A4 ≈ 27.7 cm úteis)
    COL_WIDTHS = {
        "Colaborador":            3.5 * cm,
        "Empresa":                3.5 * cm,
        "Título":                 3.5 * cm,
        "Endereço":               4.0 * cm,
        "Concluída em":           2.3 * cm,
        "Fibra Lançada (m)":      1.8 * cm,
        "Qtd CTOs":               1.4 * cm,
        "Qtd Cx Emenda":          1.4 * cm,
        "Abert./Fech. Cx Emenda": 1.8 * cm,
        "Abert./Fech. CTO":       1.5 * cm,
        "Abert./Fech. Rozeta":    1.8 * cm,
    }

    METRIC_COLS = [
        ("Qtd CTOs",               "Total de CTOs"),
        ("Qtd Cx Emenda",          "Total de Cx Emenda"),
        ("Fibra Lançada (m)",      "Total de Fibra Lançada"),
        ("Abert./Fech. Cx Emenda", "Total Ab. Cx Emenda"),
        ("Abert./Fech. CTO",       "Total Ab. CTO"),
        ("Abert./Fech. Rozeta",    "Total Ab. Rozeta"),
    ]

    def fmt_fibra(m):
        return f"{m/1000:.2f} km" if m >= 1000 else f"{m:.0f} m"

    def make_table(data_rows, col_names, widths):
        """Monta Table com Paragraphs em todas as células."""
        header_row = [Paragraph(HEADER_ALIASES.get(c, c), header_style) for c in col_names]
        body = []
        for row in data_rows:
            body.append([Paragraph(str(v), cell_style) for v in row])

        tbl = Table([header_row] + body, colWidths=widths, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND",     (0, 0), (-1, 0),  colors.HexColor("#1e3a5f")),
            ("VALIGN",         (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN",          (0, 0), (-1, 0),  "CENTER"),
            ("GRID",           (0, 0), (-1, -1), 0.4, colors.HexColor("#cccccc")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#eef2f7")]),
            ("TOPPADDING",     (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING",  (0, 0), (-1, -1), 3),
            ("LEFTPADDING",    (0, 0), (-1, -1), 3),
            ("RIGHTPADDING",   (0, 0), (-1, -1), 3),
        ]))
        return tbl

    try:
        output = io.BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=landscape(A4),
            rightMargin=1 * cm,
            leftMargin=1 * cm,
            topMargin=1 * cm,
            bottomMargin=1 * cm,
        )

        elements = []
        styles = getSampleStyleSheet()

        # ── Cabeçalho do documento ────────────────────────────────────────────
        title_style = ParagraphStyle(
            "Title", parent=styles["Heading1"], fontSize=13, alignment=1, spaceAfter=4,
        )
        sub_style = ParagraphStyle(
            "Sub", parent=styles["Normal"], fontSize=8, alignment=1, spaceAfter=3,
        )
        meses = ["", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
                 "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
        periodo = f"{meses[month]} de {year}" if month else str(year)

        elements.append(Paragraph(title, title_style))
        elements.append(Paragraph(f"Período: {periodo}", sub_style))
        elements.append(Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", sub_style))
        elements.append(Spacer(1, 8))

        # ── Preparar DataFrame — remove Observações ───────────────────────────
        df_pdf = df.drop(columns=["Observações"], errors="ignore").copy()
        for col in df_pdf.select_dtypes(include=["bool"]).columns:
            df_pdf[col] = df_pdf[col].apply(lambda x: "Sim" if x else "Não")

        # Determinar colunas e larguras na ordem do DataFrame
        cols = df_pdf.columns.tolist()
        widths = [COL_WIDTHS.get(c, 2.5 * cm) for c in cols]

        # Montar linhas
        rows = []
        for _, row in df_pdf.iterrows():
            rows.append([str(v) if v not in (None, "", "—") else "—" for v in row.values])

        elements.append(make_table(rows, cols, widths))
        elements.append(Spacer(1, 12))

        # ── Resumo Geral ──────────────────────────────────────────────────────
        section_style = ParagraphStyle(
            "Section", parent=styles["Normal"], fontSize=9,
            fontName="Helvetica-Bold", spaceAfter=3, spaceBefore=6,
        )
        sum_style = ParagraphStyle(
            "Sum", parent=styles["Normal"], fontSize=8, spaceAfter=2,
        )

        elements.append(Paragraph("Resumo Geral", section_style))
        elements.append(Paragraph(f"<b>Total de Tarefas:</b> {len(df)}", sum_style))
        for col, label in METRIC_COLS:
            if col in df.columns:
                total = df[col].sum()
                val = fmt_fibra(total) if col == "Fibra Lançada (m)" else int(total)
                elements.append(Paragraph(f"<b>{label}:</b> {val}", sum_style))

        # ── Totais por Empresa ────────────────────────────────────────────────
        if "Empresa" in df.columns:
            elements.append(Spacer(1, 8))
            elements.append(Paragraph("Tarefas por Empresa", section_style))

            metric_cols_present = [c for c, _ in METRIC_COLS if c in df.columns]
            agg: dict = {c: "sum" for c in metric_cols_present}
            # contar tarefas via coluna auxiliar
            df_count = df.copy()
            df_count["_n"] = 1
            agg["_n"] = "sum"

            grp = df_count.groupby("Empresa").agg(agg).reset_index()
            grp = grp.rename(columns={"_n": "Tarefas"})
            grp = grp.sort_values("Tarefas", ascending=False)

            if "Fibra Lançada (m)" in grp.columns:
                grp["Fibra Lançada (m)"] = grp["Fibra Lançada (m)"].apply(fmt_fibra)
            for c in metric_cols_present:
                if c != "Fibra Lançada (m)" and c in grp.columns:
                    grp[c] = grp[c].astype(int)

            emp_cols = ["Empresa", "Tarefas"] + metric_cols_present
            emp_rows = []
            for _, row in grp.iterrows():
                emp_rows.append([str(row.get(c, "—")) for c in emp_cols])

            emp_widths_map = {**COL_WIDTHS, "Empresa": 5.0 * cm, "Tarefas": 1.8 * cm}
            emp_widths = [emp_widths_map.get(c, 2.5 * cm) for c in emp_cols]

            elements.append(make_table(emp_rows, emp_cols, emp_widths))

        doc.build(elements)
        return output.getvalue()

    except Exception as e:
        print(f"Erro ao exportar PDF: {e}")
        return None
